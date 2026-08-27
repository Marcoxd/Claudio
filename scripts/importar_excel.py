#!/usr/bin/env python3
"""Importa un Excel de gastos mensual al bot.

Espera un libro con una hoja por mes (Enero, Febrero, …) y este formato:

    A: descripción       B: total de la compra (si es diferido)
    C: valor del mes     D: nombre de la tarjeta, o número de cuota
    E/F: bloque de resumen (gastos fijos, SUELDO, TARJETAS, DIFERENCIA)

Por defecto solo muestra lo que haría. Con --aplicar escribe en la base.

    python scripts/importar_excel.py Gastos.xlsx
    python scripts/importar_excel.py Gastos.xlsx --aplicar --tarjeta "Pacífico"
"""
from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import re
import unicodedata
from dataclasses import dataclass, field
from decimal import Decimal

from sqlalchemy import select

from app.db import SessionLocal, engine, init_db
from app.models import (
    ACCOUNT_CREDIT,
    KIND_EXPENSE,
    KIND_INCOME,
    STATUS_DONE,
    Account,
    Base,
    Category,
    FixedExpense,
    Person,
    RecurringIncome,
    Transaction,
)
from app.money import D, ZERO, total as suma
from app.services import buffer as buffer_service
from app.services.cards import build_installments
from app.services.fallback import guess_category
from app.services.periods import clamp_day, parse_period, period_label

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}
# Etiquetas del bloque resumen que no son gastos fijos
NO_ES_FIJO = ("TARJETA", "TOTAL", "SUELDO", "DIFERENCIA", "INGRESO")
FECHA_RE = re.compile(r"(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?")
# Un diferido se reconoce por su fórmula: =B6/3 significa "el total en 3 cuotas"
CUOTAS_RE = re.compile(r"^=\+?\s*B\d+\s*/\s*(\d+)\s*$", re.IGNORECASE)


def _plano(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", (texto or "").lower())
        if unicodedata.category(c) != "Mn"
    ).strip()


def mes_de_hoja(nombre: str) -> int | None:
    plano = _plano(nombre)
    for mes, numero in MESES.items():
        if mes in plano:
            return numero
    return None


@dataclass
class Fila:
    hoja: str
    numero: int
    descripcion: str
    valor: Decimal
    total: Decimal | None = None      # columna B: precio completo de un diferido
    cuotas: int = 1
    tarjeta: str | None = None
    fecha: dt.date | None = None


@dataclass
class Informe:
    periodos: list[str] = field(default_factory=list)
    filas: list[Fila] = field(default_factory=list)
    diferidos: list[Fila] = field(default_factory=list)
    fijos: dict[str, Decimal] = field(default_factory=dict)
    fijos_por_mes: dict[str, dict[str, Decimal]] = field(default_factory=dict)
    sueldo: Decimal = ZERO
    tarjetas: set[str] = field(default_factory=set)
    avisos: list[str] = field(default_factory=list)


def leer(ruta: str, anio: int) -> Informe:
    import openpyxl
    from openpyxl.utils import get_column_letter

    libro = openpyxl.load_workbook(ruta, data_only=True)
    formulas = openpyxl.load_workbook(ruta, data_only=False)
    informe = Informe()
    vistos: dict[str, Fila] = {}   # descripción de diferido -> primera aparición

    for nombre in libro.sheetnames:
        mes = mes_de_hoja(nombre)
        if mes is None:
            continue
        hoja = libro[nombre]
        hoja_formulas = formulas[nombre]
        periodo = f"{anio:04d}-{mes:02d}"
        informe.periodos.append(periodo)
        fijos_del_mes: dict[str, Decimal] = {}

        for fila in hoja.iter_rows():
            celdas = {
                get_column_letter(c.column): c.value
                for c in fila
                if getattr(c, "column", None)
            }
            etiqueta, b, c, d = (celdas.get(k) for k in "ABCD")

            # --- bloque de resumen (columnas E/F) ---
            e, f = celdas.get("E"), celdas.get("F")
            if isinstance(e, str) and isinstance(f, (int, float)):
                clave = e.strip().upper()
                if clave.startswith("SUELDO"):
                    informe.sueldo = max(informe.sueldo, D(f))
                elif not any(clave.startswith(x) for x in NO_ES_FIJO):
                    fijos_del_mes[e.strip()] = D(f)

            # --- gastos (columnas A–D) ---
            if isinstance(etiqueta, dt.datetime):
                fecha_fila, etiqueta = etiqueta.date(), "Gasolina"
            else:
                fecha_fila = None
            if not isinstance(etiqueta, str) or not etiqueta.strip():
                continue
            if "GASTOS TARJETAS" in etiqueta.upper():
                continue

            if isinstance(c, str) and c.strip():
                crudo = c.strip().replace(",", ".")
                try:
                    c = float(crudo)
                    informe.avisos.append(
                        f"{nombre.strip()} fila {fila[0].row}: «{etiqueta.strip()}» "
                        f"tenía {c:.2f} escrito como texto — Excel no lo sumaba."
                    )
                except ValueError:
                    continue
            if not isinstance(c, (int, float)) or not c:
                continue

            tarjeta = d.strip() if isinstance(d, str) and not d.strip().rstrip(".").isdigit() else None
            if tarjeta and not re.search(r"\d", tarjeta):
                informe.tarjetas.add(tarjeta)
            else:
                tarjeta = None

            # El diferido se detecta por la fórmula =B/N, no por la proporción:
            # así un valor que casualmente cabe N veces no se confunde con cuotas.
            cuotas = 1
            formula = hoja_formulas.cell(row=fila[0].row, column=3).value
            if isinstance(formula, str):
                encontrado = CUOTAS_RE.match(formula.strip())
                if encontrado and isinstance(b, (int, float)) and b:
                    cuotas = int(encontrado.group(1))

            fecha = fecha_fila or _fecha_en_texto(etiqueta, anio, mes)
            registro = Fila(
                hoja=nombre.strip(), numero=fila[0].row,
                descripcion=etiqueta.strip(), valor=D(c),
                total=D(b) if isinstance(b, (int, float)) and b else None,
                cuotas=cuotas, tarjeta=tarjeta,
                fecha=fecha or clamp_day(anio, mes, 15),
            )

            if cuotas > 1:
                clave = _plano(registro.descripcion)
                if clave in vistos:
                    continue          # ya lo contamos en su primer mes
                vistos[clave] = registro
                informe.diferidos.append(registro)
            else:
                informe.filas.append(registro)

        # Los fijos cambian a lo largo del año: para el histórico vale el bloque
        # de cada mes; para lo que viene, el del último.
        if fijos_del_mes:
            informe.fijos_por_mes[periodo] = fijos_del_mes
            informe.fijos = fijos_del_mes

    return informe


def _fecha_en_texto(texto: str, anio: int, mes: int) -> dt.date | None:
    encontrado = FECHA_RE.search(texto)
    if not encontrado:
        return None
    dia, mes_txt, anio_txt = encontrado.groups()
    try:
        return dt.date(
            int(anio_txt) + 2000 if anio_txt and len(anio_txt) == 2 else int(anio_txt or anio),
            int(mes_txt), int(dia),
        )
    except ValueError:
        return None


# --------------------------------------------------------------- escritura


async def aplicar(informe: Informe, tarjeta_default: str, colchon: Decimal,
                  personas: list[str], corte: int | None = None, pago: int | None = None):
    async with SessionLocal() as s:
        cats = {c.name: c for c in (await s.execute(select(Category))).scalars().all()}
        nombres_cat = list(cats)

        # --- tarjetas ---
        tarjetas: dict[str, Account] = {}
        for nombre in sorted(informe.tarjetas) or [tarjeta_default]:
            cuenta = (
                await s.execute(select(Account).where(Account.name == nombre))
            ).scalar_one_or_none()
            if cuenta is None:
                cuenta = Account(name=nombre, type=ACCOUNT_CREDIT)
                s.add(cuenta)
                await s.flush()
            if nombre == tarjeta_default:
                cuenta.cut_day = corte or cuenta.cut_day
                cuenta.due_day = pago or cuenta.due_day
            tarjetas[nombre] = cuenta
        predeterminada = tarjetas.get(tarjeta_default) or next(iter(tarjetas.values()))

        # --- personas ---
        for nombre in personas:
            existe = (
                await s.execute(select(Person).where(Person.name.ilike(nombre)))
            ).scalar_one_or_none()
            if existe is None:
                s.add(Person(name=nombre))
        await s.flush()

        # --- gastos fijos e ingreso, vigentes desde el último mes leído ---
        desde = informe.periodos[-1] if informe.periodos else None
        for nombre, monto in informe.fijos.items():
            existe = (
                await s.execute(select(FixedExpense).where(FixedExpense.name == nombre))
            ).scalar_one_or_none()
            if existe is None:
                s.add(
                    FixedExpense(
                        name=nombre, amount=monto, due_day=5,
                        category_id=cats[guess_category(nombre, nombres_cat)].id,
                        start_period=desde,
                    )
                )
        if informe.sueldo > 0 and not (await s.execute(select(RecurringIncome))).first():
            s.add(
                RecurringIncome(
                    name="Sueldo", amount=informe.sueldo, pay_day=30, start_period=desde
                )
            )
        await s.flush()

        # --- colchón ---
        if colchon > 0:
            await buffer_service.set_total(s, colchon)

        # --- fijos históricos: el mes en curso lo genera el bot, no el importador ---
        definiciones = {
            f.name: f for f in (await s.execute(select(FixedExpense))).scalars().all()
        }
        for periodo, bloque in informe.fijos_por_mes.items():
            if periodo == desde:
                continue
            anio, mes = parse_period(periodo)
            for nombre, monto in bloque.items():
                s.add(
                    Transaction(
                        kind=KIND_EXPENSE, status=STATUS_DONE,
                        date=clamp_day(anio, mes, 5), period=periodo,
                        amount=monto, description=nombre,
                        category_id=cats[guess_category(nombre, nombres_cat)].id,
                        fixed_expense_id=(
                            definiciones[nombre].id if nombre in definiciones else None
                        ),
                        source="fixed", notes=f"Gasto fijo importado ({periodo})",
                    )
                )

        # --- gastos sueltos ---
        for fila in informe.filas:
            cuenta = tarjetas.get(fila.tarjeta or "", predeterminada)
            tx = Transaction(
                kind=KIND_EXPENSE, status=STATUS_DONE, date=fila.fecha,
                period=f"{fila.fecha.year:04d}-{fila.fecha.month:02d}",
                amount=fila.valor, description=fila.descripcion[:200],
                category_id=cats[guess_category(fila.descripcion, nombres_cat)].id,
                account_id=cuenta.id, source="manual",
                notes=f"Importado de {fila.hoja} fila {fila.numero}",
            )
            s.add(tx)
            await s.flush()
            # sin cuota no aparecería en ningún estado de cuenta
            for cuota in build_installments(tx, cuenta, 1):
                s.add(cuota)

        # --- diferidos: una compra con sus cuotas ---
        for fila in informe.diferidos:
            cuenta = tarjetas.get(fila.tarjeta or "", predeterminada)
            tx = Transaction(
                kind=KIND_EXPENSE, status=STATUS_DONE, date=fila.fecha,
                period=f"{fila.fecha.year:04d}-{fila.fecha.month:02d}",
                amount=fila.total or D(fila.valor * fila.cuotas),
                description=fila.descripcion[:200],
                category_id=cats[guess_category(fila.descripcion, nombres_cat)].id,
                account_id=cuenta.id, installments_total=fila.cuotas, source="manual",
                notes=f"Importado de {fila.hoja} fila {fila.numero}",
            )
            s.add(tx)
            await s.flush()
            for cuota in build_installments(tx, cuenta, fila.cuotas):
                s.add(cuota)

        # --- sueldo de cada mes histórico ---
        if informe.sueldo > 0:
            for periodo in informe.periodos:
                anio, mes = parse_period(periodo)
                s.add(
                    Transaction(
                        kind=KIND_INCOME, status=STATUS_DONE,
                        date=clamp_day(anio, mes, 30), period=periodo,
                        amount=informe.sueldo, description="Sueldo",
                        income_type="sueldo", source="manual",
                        category_id=cats["Sueldo"].id,
                    )
                )
        await s.commit()


def mostrar(informe: Informe, tarjeta_default: str, colchon: Decimal, personas: list[str]):
    print("═" * 66)
    print("  LO QUE ENCONTRÉ")
    print("═" * 66)
    print(f"Meses: {len(informe.periodos)} "
          f"({period_label(informe.periodos[0])} → {period_label(informe.periodos[-1])})")
    print(f"Gastos sueltos: {len(informe.filas)} por {suma(f.valor for f in informe.filas)}")
    print(f"Compras a cuotas: {len(informe.diferidos)}")
    for f in informe.diferidos:
        print(f"   · {f.descripcion[:26]:<28} {f.total} en {f.cuotas} de {f.valor} "
              f"(desde {f.hoja})")
    print(f"\nGastos fijos vigentes ({suma(informe.fijos.values())} al mes):")
    for nombre, monto in informe.fijos.items():
        print(f"   · {nombre:<20} {monto}")
    print(f"\nSueldo: {informe.sueldo}")
    print(f"Tarjetas: {', '.join(sorted(informe.tarjetas)) or '(ninguna en la hoja)'}"
          f"   → por defecto: {tarjeta_default}")
    print(f"Personas: {', '.join(personas) or '(ninguna)'}")
    print(f"Colchón: {colchon}")

    print("\n" + "─" * 66)
    print("  QUÉ TAN FIEL VA A SER EL HISTÓRICO")
    print("─" * 66)
    print("  · La hoja no guarda la fecha de cada compra ni el día de corte de")
    print("    las tarjetas, así que los meses anteriores quedan aproximados:")
    print("    sirven para ver categorías y tendencias, no para cuadrar al centavo.")
    print("  · Las compras a cuotas se cargan completas el mes que empezaron,")
    print("    y el bot reparte las cuotas: por eso un mes viejo puede verse")
    print("    distinto a tu hoja.")
    print("  · De aquí en adelante sí es exacto, porque cada gasto entra con su")
    print("    fecha real y su tarjeta.")

    if informe.avisos:
        print("\n" + "─" * 66)
        print("  ERRORES QUE TRAÍA LA HOJA")
        print("─" * 66)
        for aviso in informe.avisos:
            print(f"  ⚠ {aviso}")


async def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("archivo")
    parser.add_argument("--anio", type=int, default=dt.date.today().year)
    parser.add_argument("--tarjeta", default="Tarjeta",
                        help="tarjeta a la que asignar lo que no diga cuál")
    parser.add_argument("--corte", type=int, default=None,
                        help="día de corte de esa tarjeta (mira tu estado de cuenta)")
    parser.add_argument("--pago", type=int, default=None,
                        help="día máximo de pago de esa tarjeta")
    parser.add_argument("--colchon", type=Decimal, default=Decimal(0),
                        help="dinero ajeno que administras")
    parser.add_argument("--personas", default="",
                        help="nombres separados por coma con quienes compartes gastos")
    parser.add_argument("--aplicar", action="store_true", help="escribir en la base")
    parser.add_argument("--reset", action="store_true", help="borrar todo antes")
    args = parser.parse_args()

    informe = leer(args.archivo, args.anio)
    personas = [p.strip() for p in args.personas.split(",") if p.strip()]
    mostrar(informe, args.tarjeta, D(args.colchon), personas)

    if not args.aplicar:
        print("\nSimulación: no escribí nada. Agrega --aplicar para cargarlo.")
        return

    if args.reset:
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.drop_all)
    await init_db()
    await aplicar(informe, args.tarjeta, D(args.colchon), personas,
                  args.corte, args.pago)
    print("\n✔ Cargado. Revisa /resumen y /tarjetas en el bot.")
    faltan = [t for t in sorted(informe.tarjetas) if t != args.tarjeta or not args.corte]
    if faltan:
        print("  Falta el día de corte y de pago de: " + ", ".join(faltan))
        print("  Están en la primera página del estado de cuenta, como")
        print("  «Fecha de corte» y «Fecha máxima de pago sin recargos».")
        print("  Ponlos con /tarjetas → Editar fechas.")


if __name__ == "__main__":
    asyncio.run(main())
